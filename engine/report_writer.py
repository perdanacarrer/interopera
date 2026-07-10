"""report_writer.py — Populates the report_template.xlsx with computed figures."""
import os, shutil
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
TEMPLATE = os.path.join(BASE_DIR, "sample_docs", "report_template.xlsx")
REPORTS_DIR = os.path.join(BASE_DIR, "reports")

METRIC_TO_FIGURE = {
    "Singapore Government Securities":   "allocation_Singapore Government Securities",
    "MAS Bills":                          "allocation_MAS Bills",
    "Investment Grade Corporate Bonds":   "allocation_Investment Grade Corporate Bonds",
    "High Yield Bonds":                   "allocation_High Yield Bonds",
    "Foreign Currency Bonds (hedged)":    "allocation_Foreign Currency Bonds",
    "Structured Credit (ABS/MBS)":        "allocation_Structured Credit",
    "Cash & Cash Equivalents":            "allocation_Cash & Cash Equivalents",
    "Aggregate non-IG exposure":          "aggregate_non_ig_exposure",
    "Largest single corporate issuer":    "largest_single_corporate_issuer",
    "Largest GRE issuer":                 "largest_gre_issuer",
    "Liquid assets ratio":                "liquid_assets_ratio",
    "Portfolio modified duration":        "portfolio_modified_duration",
    "Portfolio DV01":                     "portfolio_dv01",
}

def write_report(figures, firm_id, narrative=""):
    os.makedirs(REPORTS_DIR, exist_ok=True)
    out = os.path.join(REPORTS_DIR, f"{firm_id}_report.xlsx")
    shutil.copy(TEMPLATE, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active

    fig_by_key = {f["figure"]: f for f in figures}
    for row in ws.iter_rows(min_row=2):
        metric = str(row[1].value or "").strip()
        fig = fig_by_key.get(METRIC_TO_FIGURE.get(metric, ""))
        if not fig: continue
        row[2].value = fig.get("value")
        row[3].value = fig.get("limit")
        row[4].value = fig.get("utilization")
        row[5].value = fig.get("status")
        cit = fig.get("citation", {})
        row[6].value = (f"{fig.get('graph_path','')} | "
                        f"{cit.get('source_doc','')} p{cit.get('page','')} [{cit.get('chunk_id','')}]")

    # Only write Narrative sheet when we have real narrative text
    if narrative and narrative.strip() and not narrative.startswith("["):
        if "Narrative" in wb.sheetnames:
            del wb["Narrative"]
        ns = wb.create_sheet("Narrative")
        ns["A1"] = f"Compliance Commentary — {firm_id} (LLM-generated, Gemini)"
        ns["A1"].font = openpyxl.styles.Font(bold=True)
        ns["A2"] = narrative
        ns.column_dimensions["A"].width = 120

    wb.save(out)
    return out
